from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .utils import parse_amount


_SBER_DATE_RE = re.compile(r"(\d{2}\.\d{2}\.\d{4})\s+(\d{2}\.\d{2}\.\d{4})")


def read_first_sheet(path: str | Path) -> pd.DataFrame:
    return pd.read_excel(path)


def read_bank_statement(path: str | Path) -> pd.DataFrame:
    """
    Reads the uploaded Sber statement format where rows are repeated across pages,
    headers are not tabular, and real operations look like:
    A='12.06.2025 12.06.2025', B='02:00', D='Перевод...', E='+2 750,00'
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    records: list[dict] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = list(row) + [None] * 5
        a, b, c, d, e = values[:5]
        if not isinstance(a, str):
            continue
        if e is None:
            continue

        match = _SBER_DATE_RE.search(a.strip())
        amount = parse_amount(e)
        if not match or amount is None:
            continue

        op_date = pd.to_datetime(match.group(1), dayfirst=True, errors="coerce")
        if pd.isna(op_date):
            continue

        records.append(
            {
                "row_number": row_idx,
                "date_raw": a,
                "date_parsed": op_date.date(),
                "time_raw": b,
                "category": d,
                "amount_raw": e,
                "amount_parsed": amount,
            }
        )

    if records:
        return pd.DataFrame(records)

    # Fallback for more standard bank exports
    df = pd.read_excel(path)
    return df


def write_report(report: pd.DataFrame, report_path: str | Path) -> Path:
    out = Path(report_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        report.to_excel(writer, sheet_name="report", index=False)
        ws = writer.book["report"]

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for idx, col_name in enumerate(report.columns, start=1):
            width = max(len(str(col_name)), 14)
            if col_name in {"Название гаража", "Статус"}:
                width = max(width, 18)
            elif "дата" in col_name.lower():
                width = max(width, 18)
            elif "сумма" in col_name.lower():
                width = max(width, 14)

            ws.column_dimensions[get_column_letter(idx)].width = width

        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                header = ws.cell(1, col).value
                cell = ws.cell(row, col)
                if header and "дата" in str(header).lower() and cell.value:
                    cell.number_format = "DD.MM.YYYY"
                if header and "сумма" in str(header).lower() and cell.value is not None:
                    cell.number_format = '#,##0.00'

    return out
